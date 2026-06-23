import sys
import os
import re
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

thin_border_side = Side(style='thin', color='A0A0A0')
header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

# Reconfigure stdout/stderr to support UTF-8 encoding (prevents UnicodeEncodeError on Windows)
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

# ══════════════════════════════════════════════════════════════════
#  CONSOLIDATED STOCK SUMMARY — PARENT ROW NAMES
# ══════════════════════════════════════════════════════════════════
ALL_STOCK_PARENT_ROWS = {
    # Top-level & Sonepat
    'Finished Goods', 'New Products', 'RM PM', 'PACKING MATERIAL - PM',
    'Packing Material', 'Agrifort Export Bags', 'HDPE BAGS', 'Polythene Bags',
    'Pouch', 'Packing Material(1)', 'BUCKETS', 'RAW MATERIAL-RM',
    'Semi Finished Goods', 'Grand Total',
    # Sonepat Mid-level
    'FINISH GOODS BULK', 'Haryali DF (BRAND)', 'Gain', 'Insta Bor 150',
    'Insta Cal 160', 'Insta Sul 800', 'Nutrizin 700', 'Ziddi DF',
    'Profit DF', 'Sulton G', 'Super 90%', 'UPL', 'DRUMS',
    'OUTER BAGS 30KG', 'OUTER BOX', 'Paper Bags', 'LABELS/ STICKERS',
    'Insulf stickers', 'LEAFLETS', 'Chemicals', 'OTHER INGREDIENTS (BOM)',
    'DN', 'JSC', 'Ligno', 'Sulphur', 'Jaisulf DF', 'Crude Sulphur', 'Zircosil Micro Beads',
    # New Top-level
    'Finishied Good', 'Stock at Branch', 'WDG Total', 'Lab Chemicals',
    # New Finished Goods mid-level / Raw Material mid-level / Packing Material mid-level
    'New  Products RM', 'Gain RM', 'Insta Bor -RM', 'Insta Cal -RM', 'Nutrizin- RM',
    'Ziddi DF- RM', 'Sulphur Powder -RM', 'Sulphur Products', 'Sulphur WDG- RM',
    'Colour', 'FZU1', 'Other Ingredients', 'New Products PM', 'New Product Caps',
    'New Products Bottles', 'New Products Boxes', 'New Products Labels',
    # Categories from Stock Summary New
    'ELECTRICAL ITEM', 'Machinery Accessories', 'Packing Material (Nos)',
    # Extra Categories
    'Chemicals for New Products'
}

# ══════════════════════════════════════════════════════════════════
#  FILE 3 — Debtors file rows to remove
#  Filename must contain: "debtor"
#  ADD any new Jaishil branch names here in future
# ══════════════════════════════════════════════════════════════════
ROWS_TO_REMOVE_DEBTORS = {
    'Jaishil Sulphur & Chemical Ind.-Bathinda',
    'Jaishil Sulphur & Chemical Indus- Chhattishgarh',
    'Jaishil Sulphur & Chemical Industries (Akola)',
    'Jaishil Sulphur & Chemical Industries-Gujarat',
    'Jaishil Sulphur & Chemical Industries -Hissar',
    'Jaishil Sulphur & Chemical Industries- Jaipur (Raj)',
    'Jaishil Sulphur & Chemical Industries - Karnal',
    'Jaishil Sulphur & Chemical Industries (Karnataka)',
    'Jaishil Sulphur & Chemical Industries - (Lucknow)',
    'Jaishil Sulphur & Chemical Industries (Meerut)',
    'Jaishil Sulphur & Chemical Industries ( M.P)',
    'Jaishil Sulphur & Chemical Industries - PUNE',
    'Jaishil Sulphur & Chemical Industries - Shahabad',
    'Jaishil Sulphur & Chemical Industries - Sonepat.Haryana',
    'Jaishil Sulphur & Chemical Industries  - U.P',
}

# ══════════════════════════════════════════════════════════════════
#  PROFILE MAP — keyword in filename → (rows_to_remove, header_rows_to_skip, file_type)
#  file_type: "stock" or "debtors" (controls how cleaning works)
# ══════════════════════════════════════════════════════════════════
FILE_PROFILES = {
    'stock':   (ALL_STOCK_PARENT_ROWS, 12, 'stock'),
    'debtor':  (ROWS_TO_REMOVE_DEBTORS, 0, 'debtors'),
}


def detect_profile(filename):
    name_lower = os.path.basename(filename).lower()
    # Prioritize debtor profile to avoid collision with 'new' in debtors file name
    if 'debtor' in name_lower:
        return 'debtor', FILE_PROFILES['debtor']
    # Match any generic stock terms (sonepat, new, stock, summary, punjab, bathinda)
    for kw in ('sonepat', 'new', 'stock', 'summary', 'punjab', 'bathinda'):
        if kw in name_lower:
            return 'stock', FILE_PROFILES['stock']
    for keyword, profile in FILE_PROFILES.items():
        if keyword in name_lower:
            return keyword, profile
    return None, None


# ──────────────────────────────────────────────
#  STOCK FILE CLEANER (Sonepat / New)
# ──────────────────────────────────────────────
def clean_stock_file(ws, parent_rows, header_rows_to_skip):
    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))

    # Find Particulars row and column index dynamically
    particulars_row_idx = 0
    particulars_col_idx = 0
    for r_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
        for col_idx, val in enumerate(row):
            if val and str(val).strip().lower() in ('particulars', 'particular'):
                particulars_row_idx = r_idx
                particulars_col_idx = col_idx
                break
        else:
            continue
        break

    if particulars_row_idx == 0:
        particulars_row_idx = header_rows_to_skip if header_rows_to_skip > 0 else 12

    # Find Subheader row index (containing Quantity/Rate/Value) starting from particulars_row_idx
    subheader_row_idx = particulars_row_idx
    for r_idx in range(particulars_row_idx, min(particulars_row_idx + 6, ws.max_row + 1)):
        row_vals = [ws.cell(row=r_idx, column=c).value for c in range(1, ws.max_column + 1)]
        row_vals_lower = [str(val).strip().lower() for val in row_vals if val is not None]
        if any(kw in row_vals_lower for kw in ('quantity', 'rate', 'value')):
            subheader_row_idx = r_idx
            break

    # Check if there's a section row above the subheader row
    section_row_idx = None
    if subheader_row_idx > 1:
        section_keywords = {'balance', 'inward', 'outward', 'opening', 'closing'}
        has_section_keywords = False
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=subheader_row_idx - 1, column=col).value
            if val:
                val_lower = str(val).strip().lower()
                if any(kw in val_lower for kw in section_keywords):
                    has_section_keywords = True
                    break
        if has_section_keywords:
            section_row_idx = subheader_row_idx - 1

    # Extract dynamic headers before row deletion
    headers = []
    current_section = None
    max_cols = ws.max_column

    section_row_vals = []
    if section_row_idx:
        section_row_vals = [ws.cell(row=section_row_idx, column=col).value for col in range(1, max_cols + 1)]
    
    # Read from subheader row
    particulars_row_vals = [ws.cell(row=subheader_row_idx, column=col).value for col in range(1, max_cols + 1)]
    # Override the particulars column with the actual particulars cell value
    if particulars_col_idx < len(particulars_row_vals):
        particulars_row_vals[particulars_col_idx] = ws.cell(row=particulars_row_idx, column=particulars_col_idx + 1).value

    for col_idx in range(max_cols):
        sec_val = section_row_vals[col_idx] if section_row_idx and col_idx < len(section_row_vals) else None
        part_val = particulars_row_vals[col_idx] if col_idx < len(particulars_row_vals) else None
        
        if sec_val is not None and str(sec_val).strip() != '':
            current_section = str(sec_val).strip()
            
        part_str = str(part_val).strip() if part_val is not None else ''
        
        if part_str.lower() in ('particulars', 'particular'):
            headers.append(part_str if part_str else 'Particulars')
        else:
            if current_section and part_str:
                headers.append(f"{current_section} - {part_str}")
            elif part_str:
                headers.append(part_str)
            elif current_section:
                headers.append(current_section)
            else:
                headers.append('')

    # Check if column immediately after particulars column is empty in headers but has data
    has_warehouse = False
    if particulars_col_idx + 1 < len(headers) and not headers[particulars_col_idx + 1]:
        has_data = False
        for r in range(subheader_row_idx + 1, min(subheader_row_idx + 21, ws.max_row + 1)):
            cell_val = ws.cell(row=r, column=particulars_col_idx + 2).value
            if cell_val is not None and str(cell_val).strip() != '':
                has_data = True
                break
        if has_data:
            # Delete the Warehouse/Batch column entirely since it only contains data on rows that get deleted
            ws.delete_cols(particulars_col_idx + 2)
            if particulars_col_idx + 1 < len(headers):
                headers.pop(particulars_col_idx + 1)
            if particulars_col_idx + 1 < len(particulars_row_vals):
                particulars_row_vals.pop(particulars_col_idx + 1)
            if section_row_idx and particulars_col_idx + 1 < len(section_row_vals):
                section_row_vals.pop(particulars_col_idx + 1)
            max_cols -= 1
            has_warehouse = False

    all_rows = list(ws.iter_rows(values_only=False))
    rows_to_delete = []
    change_log = []

    normalized_parents = {str(name).strip().lower() for name in parent_rows}

    # Mark header rows
    for idx in range(1, subheader_row_idx + 1):
        if idx <= len(all_rows):
            rows_to_delete.append(idx)
            row = all_rows[idx-1]
            val_a = row[0].value
            val_b = row[1].value if len(row) > 1 else None
            val_c = row[2].value if len(row) > 2 else None
            val_d = row[3].value if len(row) > 3 else None
            change_log.append((idx, val_a, val_b, val_c, val_d, 'Header/Title Row'))

    # Process data rows
    for idx in range(subheader_row_idx + 1, len(all_rows) + 1):
        row = all_rows[idx-1]
        if particulars_col_idx >= len(row):
            continue
            
        cell_a = row[particulars_col_idx]
        val_a = cell_a.value if cell_a else None
        
        val_b = row[particulars_col_idx + 1].value if particulars_col_idx + 1 < len(row) else None
        val_c = row[particulars_col_idx + 2].value if particulars_col_idx + 2 < len(row) else None
        val_d = row[particulars_col_idx + 3].value if particulars_col_idx + 3 < len(row) else None

        if val_a is None or str(val_a).strip() == '':
            rows_to_delete.append(idx)
            change_log.append((idx, val_a, val_b, val_c, val_d, 'Empty Row'))
            continue

        val_a_str = str(val_a).strip()
        val_a_lower = val_a_str.lower()
        
        if 'grand total' in val_a_lower:
            rows_to_delete.append(idx)
            change_log.append((idx, val_a, val_b, val_c, val_d, 'Grand Total Row'))
            continue

        if any(kw in val_a_lower for kw in ('brochure', 'broucher', 'banner', 'sticker', 'outer bag')):
            rows_to_delete.append(idx)
            change_log.append((idx, val_a, val_b, val_c, val_d, 'Non-product row removed'))
            continue

        if 'primary batch' in val_a_lower:
            rows_to_delete.append(idx)
            change_log.append((idx, val_a, val_b, val_c, val_d, 'Primary Batch Row Removed'))
            continue

        is_parent = False
        
        # Determine where numeric columns start
        start_numeric_idx = particulars_col_idx + 1
        if has_warehouse:
            start_numeric_idx = particulars_col_idx + 2
            
        # Check if all numeric columns are empty
        numeric_vals = [row[c].value for c in range(start_numeric_idx, len(row))]
        all_numeric_empty = all(v is None or str(v).strip() == '' for v in numeric_vals)

        # 1. Check name list
        if val_a_lower in normalized_parents:
            is_parent = True
        # 2. Check bold font
        elif cell_a.font and cell_a.font.bold:
            is_parent = True
        # 3. Check if all numeric columns are empty
        elif all_numeric_empty:
            is_parent = True
        else:
            # 4. Check stock branch category row
            val_a_clean = normalize_text(val_a_str)
            if val_a_clean.startswith("stock at "):
                branch_part = val_a_clean[len("stock at "):].strip().rstrip('.').strip()
                if branch_part in {
                    'bathinda', 'chhattishgarh', 'chhatisgarh', 'chatisgarh', 'akola', 
                    'gujarat', 'hissar', 'hisar', 'jaipur', 'karnal', 'karnataka', 
                    'lucknow', 'meerut', 'm p', 'pune', 'shahabad', 'sonepat', 
                    'u p'
                }:
                    is_parent = True

            # 5. Check dynamic look-ahead parent row detection
            if not is_parent and val_a_clean:
                child_rows = []
                for k in range(idx, len(all_rows)):
                    next_row = all_rows[k]
                    if particulars_col_idx >= len(next_row):
                        break
                    next_cell_a = next_row[particulars_col_idx]
                    next_val_a = next_cell_a.value if next_cell_a else None
                    if next_val_a is None or str(next_val_a).strip() == "":
                        continue
                    
                    next_val_a_str = str(next_val_a).strip()
                    next_val_a_lower = next_val_a_str.lower()
                    if 'primary batch' in next_val_a_lower or 'brochure' in next_val_a_lower or 'broucher' in next_val_a_lower:
                        continue

                    next_val_a_norm = normalize_text(next_val_a_str)
                    if next_val_a_norm.startswith(val_a_clean) and next_val_a_norm != val_a_clean:
                        child_rows.append(next_row)
                    else:
                        break
                
                if child_rows:
                    parent_numeric = []
                    for col_idx in range(start_numeric_idx, len(row)):
                        val = row[col_idx].value
                        try:
                            parent_numeric.append(float(val) if val is not None else 0.0)
                        except (ValueError, TypeError):
                            parent_numeric.append(0.0)
                    
                    child_sums = [0.0] * len(parent_numeric)
                    for child in child_rows:
                        for col_idx in range(start_numeric_idx, len(child)):
                            val = child[col_idx].value
                            try:
                                f_val = float(val) if val is not None else 0.0
                                c_idx = col_idx - start_numeric_idx
                                if c_idx < len(child_sums):
                                    child_sums[c_idx] += f_val
                            except (ValueError, TypeError):
                                pass
                    
                    for p_val, c_sum in zip(parent_numeric, child_sums):
                        if p_val != 0.0 and abs(p_val - c_sum) < 1.0:
                            is_parent = True
                            break

        if is_parent:
            rows_to_delete.append(idx)
            change_log.append((idx, val_a, val_b, val_c, val_d, 'Parent/Subtotal Row'))

    for row_num in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_num)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", start_color="D9E1F2")
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Add clean header
    if section_row_idx is not None:
        # Build 2-row headers
        section_headers = [None] * max_cols
        subheader_headers = [None] * max_cols
        
        current_sec = None
        for c_idx in range(max_cols):
            # Particulars column
            if c_idx == particulars_col_idx:
                section_headers[c_idx] = particulars_row_vals[particulars_col_idx] or 'Particulars'
                subheader_headers[c_idx] = None
                continue
            # Warehouse/Batch column
            if has_warehouse and c_idx == particulars_col_idx + 1:
                section_headers[c_idx] = 'Warehouse/Batch'
                subheader_headers[c_idx] = None
                continue
                
            # Read section value
            sec_val = section_row_vals[c_idx] if c_idx < len(section_row_vals) else None
            if sec_val is not None and str(sec_val).strip() != '':
                current_sec = str(sec_val).strip()
            section_headers[c_idx] = current_sec
            
            # Read subheader value
            part_val = particulars_row_vals[c_idx] if c_idx < len(particulars_row_vals) else None
            subheader_headers[c_idx] = str(part_val).strip() if (part_val is not None and str(part_val).strip() != '') else None

        ws.insert_rows(1, 2)
        
        # Write values to Row 1 and Row 2
        for c_idx in range(max_cols):
            cell_r1 = ws.cell(row=1, column=c_idx + 1)
            cell_r1.value = section_headers[c_idx]
            
            cell_r2 = ws.cell(row=2, column=c_idx + 1)
            cell_r2.value = subheader_headers[c_idx]

        # Apply styling to all cells in Row 1 and Row 2
        for r in (1, 2):
            for c in range(1, max_cols + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = header_border

        # Perform vertical merges where subheader is None
        for c_idx in range(max_cols):
            if subheader_headers[c_idx] is None:
                ws.merge_cells(start_row=1, start_column=c_idx + 1, end_row=2, end_column=c_idx + 1)

        # Perform horizontal merges for sections in Row 1
        start_col = None
        current_sec_val = None
        for c_idx in range(max_cols):
            sec_val = section_headers[c_idx]
            if sec_val in ("Particulars", "Warehouse/Batch") or sec_val is None or str(sec_val).strip() == "":
                if start_col is not None and start_col < c_idx:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=c_idx)
                start_col = None
                current_sec_val = None
                continue
                
            if sec_val != current_sec_val:
                if start_col is not None and start_col < c_idx:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=c_idx)
                start_col = c_idx + 1
                current_sec_val = sec_val
                
        if start_col is not None and start_col < max_cols:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=max_cols)

    else:
        # Build single-row header
        ws.insert_rows(1, 1)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = header_border

    # Adjust column widths dynamically
    ws.column_dimensions['A'].width = 55
    for col_idx in range(2, max_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    # Ensure grid lines are visible
    if ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = True
    else:
        ws.sheet_view.showGridLines = True

    parent_removed = len(rows_to_delete) - subheader_row_idx
    return rows_to_delete, change_log, subheader_row_idx, parent_removed



def normalize_text(text):
    if not text:
        return ""
    text = str(text).lower()
    # Replace non-alphanumeric (excluding &) with spaces
    text = re.sub(r'[^a-z0-9&]', ' ', text)
    # Collapse multiple spaces
    return ' '.join(text.split())

def format_date_val(val):
    if val is None or str(val).strip() == '':
        return val
    # If it is a datetime, date, or Timestamp (using string representation check or isinstance)
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime('%Y-%m-%d')
    # If it is a pandas Timestamp-like object (avoiding direct pandas import dependancy)
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    val_str = str(val).strip()
    if ' 00:00:00' in val_str:
        return val_str.replace(' 00:00:00', '')
    if ' 0:00:00' in val_str:
        return val_str.replace(' 0:00:00', '')
    match = re.match(r'^(\d{4}-\d{2}-\d{2})\s+\d{1,2}:\d{2}:\d{2}$', val_str)
    if match:
        return match.group(1)
    return val


# ──────────────────────────────────────────────
#  DEBTORS FILE CLEANER
# ──────────────────────────────────────────────
def clean_debtors_file(ws, rows_to_remove):
    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))

    # Find Particulars row and column index dynamically (checking first 20 rows)
    particulars_row_idx = 0
    particulars_col_idx = 0  # Default to index 0
    for r_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
        for col_idx, val in enumerate(row):
            if val and str(val).strip().lower() in ('particulars', 'particular'):
                particulars_row_idx = r_idx
                particulars_col_idx = col_idx
                break
        else:
            continue
        break

    if particulars_row_idx == 0:
        particulars_row_idx = 1

    rows_to_delete = []
    change_log = []

    all_rows = list(ws.iter_rows(values_only=False))

    # Mark header metadata rows above the Particulars row
    for idx in range(1, particulars_row_idx):
        if idx <= len(all_rows):
            rows_to_delete.append(idx)
            row = all_rows[idx-1]
            val_0 = row[0].value if len(row) > 0 else None
            val_1 = row[1].value if len(row) > 1 else None
            val_2 = row[2].value if len(row) > 2 else None
            val_3 = row[3].value if len(row) > 3 else None
            change_log.append((idx, val_0, val_1, val_2, val_3, 'Header/Title Row'))

    # Cache normalized target names
    normalized_targets = [normalize_text(name) for name in rows_to_remove]

    for i in range(particulars_row_idx + 1, len(all_rows) + 1):
        row = all_rows[i-1]
        if particulars_col_idx >= len(row):
            continue
            
        cell_val = row[particulars_col_idx].value
        col_val_raw = str(cell_val) if cell_val is not None else ""
        col_val_clean = normalize_text(col_val_raw)
        
        # Check if row is completely empty
        row_vals = [cell.value for cell in row]
        is_empty = all(v is None or str(v).strip() == '' for v in row_vals)
        
        is_match = False
        reason = 'Jaishil Branch Row Removed'
        if is_empty:
            is_match = True
            reason = 'Empty Row'
        elif col_val_clean:
            # Match any row starting with 'jaishil' or present in the ROWS_TO_REMOVE_DEBTORS list
            if col_val_clean.startswith('jaishil'):
                is_match = True
            else:
                for target_clean in normalized_targets:
                    if target_clean == col_val_clean or target_clean in col_val_clean:
                        is_match = True
                        break
        
        if is_match:
            rows_to_delete.append(i)
            particular_val = row[particulars_col_idx].value
            other_vals = [cell.value for idx, cell in enumerate(row) if idx != particulars_col_idx]
            val_1 = other_vals[0] if len(other_vals) > 0 else None
            val_2 = other_vals[1] if len(other_vals) > 1 else None
            val_3 = other_vals[2] if len(other_vals) > 2 else None
            change_log.append((i, particular_val, val_1, val_2, val_3, reason))

    for row_num in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_num)

    # Style the header row (now Row 1)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", start_color="D9E1F2")
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    max_cols = ws.max_column
    for c in range(1, max_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Adjust column widths dynamically based on headers
    for col_idx in range(1, max_cols + 1):
        col_letter = get_column_letter(col_idx)
        val = ws.cell(row=1, column=col_idx).value
        val_str = str(val).lower() if val else ''
        if any(k in val_str for k in ('particular', 'material', 'customer', 'supplier', 'party', 'ledger', 'company', 'name')):
            ws.column_dimensions[col_letter].width = 55
        else:
            ws.column_dimensions[col_letter].width = 20

    # Ensure grid lines are visible
    if ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = True
    else:
        ws.sheet_view.showGridLines = True

    header_removed = particulars_row_idx - 1
    other_removed = len(rows_to_delete) - header_removed
    return rows_to_delete, change_log, header_removed, other_removed


# ──────────────────────────────────────────────
#  MAIN ENTRY POINT
# ──────────────────────────────────────────────
def process_file(input_path, output_path=None, profile_key=None):
    # Auto-detect or use forced profile
    detected_key, profile = detect_profile(input_path)

    if profile_key:
        profile_key = profile_key.lower()
        if profile_key in ('sonepat', 'new'):
            profile_key = 'stock'
        if profile_key not in FILE_PROFILES:
            print(f"❌ Unknown profile '{profile_key}'. Choose from: {list(FILE_PROFILES.keys())}")
            sys.exit(1)
        profile = FILE_PROFILES[profile_key]
        detected_key = profile_key

    if profile is None:
        print(f"\n❌ Could not auto-detect file type from: '{os.path.basename(input_path)}'")
        print(f"   Filename must contain one of: {list(FILE_PROFILES.keys())} or stock keywords")
        print(f"   Or pass manually: --stock / --debtor")
        sys.exit(1)

    rows_to_remove, header_rows_to_skip, file_type = profile

    # Output path
    if output_path is None:
        folder = os.path.dirname(os.path.abspath(input_path))
        basename = os.path.basename(input_path)
        base = basename
        for ext in ('.xlsx', '.xls'):
            if basename.lower().endswith(ext):
                base = basename[:-len(ext)]
                break
        output_path = os.path.join(folder, f"{base}_Cleaned.xlsx")

    print(f"\n📂 Input   : {input_path}")
    print(f"🔍 Profile : {detected_key.upper()} — {file_type} file ({len(rows_to_remove)} rows defined for removal)")
    print(f"💾 Output  : {output_path}")

    try:
        wb = load_workbook(input_path)
    except Exception as e:
        # Try loading as a file-like object to bypass string extension checks in openpyxl
        try:
            with open(input_path, 'rb') as f:
                wb = load_workbook(f)
        except Exception:
            raise e
    ws = wb.active

    # Run the appropriate cleaner
    if file_type == 'stock':
        rows_to_delete, change_log, header_removed, parent_removed = clean_stock_file(
            ws, rows_to_remove, header_rows_to_skip
        )
    else:
        rows_to_delete, change_log, header_removed, parent_removed = clean_debtors_file(
            ws, rows_to_remove
        )

    # Create Change Log sheet
    if 'Change Log' in wb.sheetnames:
        del wb['Change Log']
    cl = wb.create_sheet('Change Log')

    cl_headers = ['Original Row Number', 'Particulars', 'Column B', 'Column C', 'Column D', 'Reason for Removal']
    cl_widths   = [20, 60, 15, 15, 15, 30]
    for col_idx, (h, w) in enumerate(zip(cl_headers, cl_widths), start=1):
        cell = cl.cell(row=1, column=col_idx)
        cell.value = h
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill("solid", start_color="FFE699")
        cell.alignment = Alignment(horizontal='center')
        cell.border = header_border
        cl.column_dimensions[get_column_letter(col_idx)].width = w

    for row_idx, entry in enumerate(change_log, start=2):
        for col_idx, val in enumerate(entry, start=1):
            cl.cell(row=row_idx, column=col_idx).value = val

    # Ensure grid lines are visible on Change Log sheet
    if cl.views.sheetView:
        cl.views.sheetView[0].showGridLines = True
    else:
        cl.sheet_view.showGridLines = True

    # Format all date-time cells to remove the time component
    for sheet in wb.worksheets:
        if sheet.title != 'Change Log':
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.value = format_date_val(cell.value)

    wb.save(output_path)

    rows_kept = ws.max_row - 1
    print(f"\n✅ Done!")
    if file_type == 'stock':
        print(f"   Rows removed : {len(rows_to_delete)}  ({header_removed} header + {parent_removed} parent rows)")
    else:
        print(f"   Rows removed : {len(rows_to_delete)}  ({header_removed} header + {parent_removed} branch/empty rows)")
    print(f"   Rows kept    : {rows_kept}")
    print(f"   Change Log   : {len(change_log)} entries logged")

    return output_path


# ══════════════════════════════════════════════════════════════════
#  HOW TO RUN
#
#  Single file (auto-detects from filename):
#    python clean_stock_all.py Stock_summary_Sonepat.xlsx
#    python clean_stock_all.py Stock_Summary_New.xlsx
#    python clean_stock_all.py Debtors_Report.xlsx
#
#  Multiple files in one command:
#    python clean_stock_all.py Stock_summary_Sonepat.xlsx Stock_Summary_New.xlsx Debtors_Report.xlsx
#
#  Force a profile if filename doesn't have the keyword:
#    python clean_stock_all.py MyFile.xlsx --sonepat
#    python clean_stock_all.py MyFile.xlsx --new
#    python clean_stock_all.py MyFile.xlsx --debtor
#
#  Custom output name (single file only):
#    python clean_stock_all.py Debtors_Report.xlsx --out Debtors_Clean.xlsx
# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    args = sys.argv[1:]

    if not args:
        print("Usage:")
        print("  python clean_stock_all.py <file1.xlsx> [file2.xlsx ...]")
        print("  python clean_stock_all.py <file.xlsx> --stock | --debtor")
        sys.exit(1)

    files = []
    force_profile = None
    custom_out = None
    i = 0
    while i < len(args):
        if args[i] in ('--sonepat', '--new', '--stock', '--debtor'):
            force_profile = args[i].lstrip('-')
            if force_profile in ('sonepat', 'new'):
                force_profile = 'stock'
        elif args[i] == '--out' and i + 1 < len(args):
            custom_out = args[i + 1]
            i += 1
        elif args[i].endswith('.xlsx') or args[i].endswith('.xls'):
            files.append(args[i])
        i += 1

    if not files:
        print("❌ No Excel files provided.")
        sys.exit(1)

    print(f"\n🚀 Processing {len(files)} file(s)...\n" + "─" * 50)
    for f in files:
        out = custom_out if (len(files) == 1 and custom_out) else None
        process_file(f, output_path=out, profile_key=force_profile)
        print("─" * 50)

    print(f"\n🎉 All {len(files)} file(s) cleaned successfully!")