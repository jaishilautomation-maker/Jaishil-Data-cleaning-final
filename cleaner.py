# ============================================================
# cleaner.py
# ============================================================
import os
import re
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# COMMON KEYWORDS
# ============================================================

NON_PRODUCT_KEYWORDS = [
    'banner',
    'brochure',
    'broucher',
    'sticker',
    'outer bag'
]

# ============================================================
# HELPER FUNCTION
# ============================================================

def blank_row(df, index):
    df.loc[index] = [None] * len(df.columns)

def format_date_val(val):
    if pd.isna(val) or str(val).strip() == '':
        return val
    if isinstance(val, (datetime.datetime, datetime.date, pd.Timestamp)):
        return val.strftime('%Y-%m-%d')
    val_str = str(val).strip()
    if ' 00:00:00' in val_str:
        return val_str.replace(' 00:00:00', '')
    if ' 0:00:00' in val_str:
        return val_str.replace(' 0:00:00', '')
    match = re.match(r'^(\d{4}-\d{2}-\d{2})\s+\d{1,2}:\d{2}:\d{2}$', val_str)
    if match:
        return match.group(1)
    return val

def parse_date_to_fy_month(date_val):
    if pd.isna(date_val) or str(date_val).strip() == '':
        return None, None
    try:
        dt = pd.to_datetime(date_val, errors='coerce')
        if pd.isna(dt):
            parts = str(date_val).split('/')
            if len(parts) == 3:
                dt = pd.to_datetime(f"{parts[0]} {parts[1]} {parts[2]}", errors='coerce')
        
        if pd.isna(dt):
            return None, None
        
        year = dt.year
        month_num = dt.month
        if month_num in (1, 2, 3):
            fy_start = year - 1
        else:
            fy_start = year
        fy_end = (fy_start + 1) % 100
        fy = f"FY{fy_start}-{fy_end:02d}"
        
        month_name = dt.strftime('%B')
        return fy, month_name
    except Exception:
        return None, None

def parse_material_details(material_name):
    if not material_name or pd.isna(material_name):
        return "", "", None, None, "Other"
    
    # Step 4 — Special Case Overrides
    name_str = str(material_name).strip()
    name_lower = name_str.lower()
    std_name = name_str
    
    if 'hariyali df-0.05 kg' in name_lower:
        std_name = 'Hariyali DF 0.5 kg (20 kg box)'
    elif 'profit df-30 kg (l) d/m' in name_lower:
        std_name = 'Profit DF 30 kg drum'
    elif 'profit df (1x30kg) loose drum' in name_lower:
        std_name = 'Profit DF 30 kg drum'
    elif 'haryali df -500gm' in name_lower:
        std_name = 'Hariyali DF 0.5 kg (20 kg box)'
    
    std_name_lower = std_name.lower()
    
    # Step 1 — Brand Detection
    brand = "Unknown"
    if 'super hariyali' in std_name_lower:
        brand = 'Super Hariyali'
    elif 'hariyali' in std_name_lower:
        brand = 'Hariyali DF'
    elif 'profit' in std_name_lower:
        brand = 'Profit DF'
    elif 'gain' in std_name_lower:
        brand = 'Gain'
    elif 'instasul' in std_name_lower:
        brand = 'InstaSul'
    elif 'instabor 150' in std_name_lower or 'instabor-150' in std_name_lower:
        brand = 'InstaBor 150'
    elif 'tashvin ultra' in std_name_lower:
        brand = 'Tashvin Ultra'
    elif 'tashvin wg' in std_name_lower:
        brand = 'Tashvin WG'
        
    # Step 2 — Quantity Extraction
    qty_val = None
    unit_val = None
    match = re.search(r'(\d+\.?\d*)\s*(kg|kgs|gm|gms|g\b|ml|litre|liter|ltr|ltrs|l\b)', std_name_lower)
    if match:
        qty_raw = match.group(1)
        unit_raw = match.group(2)
        
        try:
            qty_val = float(qty_raw)
            if qty_val.is_integer():
                qty_val = int(qty_val)
        except ValueError:
            qty_val = qty_raw
        
        if unit_raw in ('kg', 'kgs'):
            unit_val = 'kg'
        elif unit_raw in ('gm', 'gms', 'g'):
            unit_val = 'g'
        elif unit_raw == 'ml':
            unit_val = 'ml'
        elif unit_raw in ('litre', 'liter', 'ltr', 'ltrs', 'l'):
            unit_val = 'l'
            
    # Step 3 — Packaging Detection
    packaging = "Other"
    if any(k in std_name_lower for k in ('drum', 'd/m', 'drm')):
        packaging = 'Drum'
    elif any(k in std_name_lower for k in ('bucket', 'bkt')):
        packaging = 'Bucket'
    elif any(k in std_name_lower for k in ('10l', '10 l', '10ltr', '10x1')):
        packaging = '10L Box'
    elif any(k in std_name_lower for k in ('5l', '5 l', '5ltr', '5x1')):
        packaging = '5L Box'
        
    return std_name, brand, qty_val, unit_val, packaging

def normalize_supplier_name(name):
    if not name or pd.isna(name):
        return ""
    name_str = str(name).strip()
    name_lower = name_str.lower()
    
    if 'jaishil' in name_lower:
        branch = ""
        if '-' in name_str:
            parts = name_str.split('-', 1)
            branch = parts[1].strip()
        elif 'industries' in name_lower:
            idx = name_lower.find('industries')
            branch_part = name_str[idx + len('industries'):].strip()
            if branch_part:
                branch = branch_part.strip().lstrip('-').strip()
                
        if branch:
            branch_clean = branch.title()
            return f"Jaishil Sulphur & Chemical Industries - {branch_clean}"
        else:
            return "Jaishil Sulphur & Chemical Industries"
            
    return name_str

# ============================================================
# FILE TYPE DETECTION
# ============================================================

def detect_file_type(file_path):

    try:
        raw_df = pd.read_excel(file_path, header=None)

        text = ""

        for i in range(min(10, len(raw_df))):
            row = raw_df.iloc[i].astype(str).tolist()
            text += " ".join(row).lower()

        # SALES
        if "material name" in text:
            return "SALES"

        # PURCHASE
        elif "customer name" in text and "uom" in text:
            return "PURCHASE"

        # STOCK
        elif "grand total" in text:
            return "STOCK"

        # JV
        elif "partyname" in text or "narration" in text:
            return "JV"
        # HO EXP
        elif "ho" in os.path.basename(file_path).lower():
            return "HO EXP"
        # SUNDRY CREDITORS
        elif "sundry creditors" in text or "group summary" in text:
            return "SUNDRY CREDITORS"

        else:
            return "UNKNOWN"

    except:
        return "UNKNOWN"

# ============================================================
# SALES CLEANING
# ============================================================

def clean_sales_file(file_path):

    raw_df = pd.read_excel(file_path, header=None)
    cleaned_df = raw_df.copy()

    changes = []

    header_row = None

    for idx, row in raw_df.iterrows():

        vals = [
            str(x).strip().lower()
            for x in row.tolist()
            if pd.notna(x)
        ]

        if 'material name' in vals:
            header_row = idx
            break

    if header_row is None:
        return raw_df, cleaned_df, changes

    headers = raw_df.iloc[header_row].tolist()

    data_df = raw_df.iloc[header_row + 1:].copy()
    data_df.columns = headers

    material_col = None
    date_col = None

    for col in data_df.columns:
        col_str = str(col).lower()
        if 'material' in col_str:
            material_col = col
        elif 'date' in col_str or 'dt' in col_str:
            date_col = col

    if material_col is None:
        return raw_df, cleaned_df, changes

    # Get integer indices for cleaner.py operations on cleaned_df (since it lacks headers)
    material_col_idx = headers.index(material_col)
    date_col_idx = headers.index(date_col) if date_col is not None else None

    # Initialize 7 new columns
    num_cols = len(cleaned_df.columns)
    new_headers = ['SKU_Clean', 'Brand', 'Quantity_Clean', 'Unit_Clean', 'Packaging', 'FY', 'Month']
    for idx, h in enumerate(new_headers):
        cleaned_df.at[header_row, num_cols + idx] = h

    for i in data_df.index:

        material = str(data_df.at[i, material_col]).strip()
        lower_mat = material.lower()

        # REMOVE NON PRODUCTS
        if any(k in lower_mat for k in NON_PRODUCT_KEYWORDS):

            blank_row(cleaned_df, i)

            changes.append({
                'row': i + 1,
                'reason': f'Non-product row removed: {material}'
            })

            continue

        # Standardize and extract material details
        std_name, brand, qty_val, unit_val, packaging = parse_material_details(material)

        # Overwrite the original Material Name column in-place using integer index
        cleaned_df.at[i, material_col_idx] = std_name

        # Parse Invoice Date
        date_val = data_df.at[i, date_col] if date_col is not None else None
        fy, month = parse_date_to_fy_month(date_val)

        # Populate the 7 new columns
        cleaned_df.at[i, num_cols] = std_name
        cleaned_df.at[i, num_cols + 1] = brand
        cleaned_df.at[i, num_cols + 2] = qty_val
        cleaned_df.at[i, num_cols + 3] = unit_val
        cleaned_df.at[i, num_cols + 4] = packaging
        cleaned_df.at[i, num_cols + 5] = fy
        cleaned_df.at[i, num_cols + 6] = month

        if std_name != material:

            changes.append({
                'row': i + 1,
                'reason': f'Material corrected: {material} → {std_name}'
            })

    return raw_df, cleaned_df, changes

# ============================================================
# PURCHASE CLEANING
# ============================================================

def clean_purchase_file(file_path):

    raw_df = pd.read_excel(file_path, header=None)
    cleaned_df = raw_df.copy()

    changes = []

    header_row = None

    for idx, row in raw_df.iterrows():
        val_0 = str(row[0]).strip().lower() if pd.notna(row[0]) else ''
        if val_0 in ('si no.', 'si no', 'sl no.', 'sl no', 'sr no.', 'sr no', 'sr. no.', 'sr.no.', 's.no.', 's. no.', 'sno', 'sno.'):
            header_row = idx
            break

    if header_row is None:
        return raw_df, cleaned_df, changes

    headers = raw_df.iloc[header_row].tolist()

    data_df = raw_df.iloc[header_row + 1:].copy()
    data_df.columns = headers

    cols = {str(c).strip().lower(): c for c in data_df.columns}

    material_col = None
    uom_col = None
    qty_col = None
    supplier_col = None

    for c in cols:

        if 'material' in c:
            material_col = cols[c]

        elif c == 'uom':
            uom_col = cols[c]

        elif 'quantity' in c:
            qty_col = cols[c]

        elif 'customer name' in c:
            supplier_col = cols[c]

    # Get integer indices for cleaner.py operations on cleaned_df (since it lacks headers)
    material_col_idx = headers.index(material_col) if material_col else None
    uom_col_idx = headers.index(uom_col) if uom_col else None
    qty_col_idx = headers.index(qty_col) if qty_col else None
    supplier_col_idx = headers.index(supplier_col) if supplier_col else None

    for i in data_df.index:

        # REMOVE NON PRODUCTS
        if material_col:
            material = str(data_df.at[i, material_col]).strip()
            lower_mat = material.lower()
            if any(k in lower_mat for k in NON_PRODUCT_KEYWORDS):
                blank_row(cleaned_df, i)
                changes.append({
                    'row': i + 1,
                    'reason': f'Non-product row removed: {material}'
                })
                continue

        # TON TO KG
        if uom_col_idx is not None and qty_col_idx is not None:

            uom = str(data_df.at[i, uom_col]).lower()

            if 'ton' in uom:

                try:
                    qty = float(
                        str(data_df.at[i, qty_col]).replace(',', '')
                    )

                    new_qty = qty * 1000

                    cleaned_df.at[i, qty_col_idx] = new_qty
                    cleaned_df.at[i, uom_col_idx] = 'kg'

                    changes.append({
                        'row': i + 1,
                        'reason': 'Quantity converted from ton to kg'
                    })

                except:
                    pass

        # SUPPLIER TYPO / NORMALISATION
        if supplier_col_idx is not None:

            supplier = str(data_df.at[i, supplier_col])
            corrected = normalize_supplier_name(supplier)

            if corrected != supplier:

                cleaned_df.at[i, supplier_col_idx] = corrected

                changes.append({
                    'row': i + 1,
                    'reason': f'Supplier spelling/formatting corrected: {supplier} → {corrected}'
                })

    return raw_df, cleaned_df, changes

# ============================================================
# STOCK CLEANING
# ============================================================

def clean_stock_file(file_path):

    raw_df = pd.read_excel(file_path, header=None)
    cleaned_df = raw_df.copy()

    changes = []

    for idx, row in raw_df.iterrows():

        material = str(row[0]).strip().lower() \
            if pd.notna(row[0]) else ''

        qty = row[2] if len(row) > 2 else None
        value = row[4] if len(row) > 4 else None

        # REMOVE NON PRODUCTS
        if any(k in material for k in NON_PRODUCT_KEYWORDS):

            blank_row(cleaned_df, idx)

            changes.append({
                'row': idx + 1,
                'reason': 'Non-product row removed'
            })

            continue

        qty_blank = pd.isna(qty) or str(qty).strip() == ''
        value_present = pd.notna(value) and str(value).strip() != ''

        # REMOVE SUBTOTALS
        if qty_blank and value_present:

            if 'grand total' not in material:

                blank_row(cleaned_df, idx)

                changes.append({
                    'row': idx + 1,
                    'reason': 'Subtotal/category removed'
                })

    return raw_df, cleaned_df, changes

# ============================================================
# JV CLEANING
# ============================================================

def clean_jv_file(file_path):

    raw_df = pd.read_excel(file_path, header=None)
    cleaned_df = raw_df.copy()

    changes = []

    header_row = None

    for idx, row in raw_df.iterrows():

        vals = [
            str(x).strip().lower()
            for x in row.tolist()
            if pd.notna(x)
        ]

        if any(term in vals for term in (
            'si no.', 'si no', 'sl no.', 'sl no', 'sr no.', 'sr no', 'sr. no.', 'sr.no.', 
            's.no.', 's. no.', 'sno', 'sno.', 'particulars', 'partyname', 'party name', 'narration'
        )):
            header_row = idx
            break

    if header_row is None:
        return raw_df, cleaned_df, changes

    headers = raw_df.iloc[header_row].tolist()

    data_df = raw_df.iloc[header_row + 1:].copy()
    data_df.columns = headers

    cols = {str(c).strip().lower(): c for c in data_df.columns}

    party_col = None
    particulars_col = None
    narration_col = None
    credit_col = None

    for c in cols:
        c_clean = c.replace(' ', '').replace('_', '').replace('.', '')

        if 'partyname' in c_clean or 'party' in c_clean or 'ledger' in c_clean:
            party_col = cols[c]

        elif 'particulars' in c_clean:
            particulars_col = cols[c]

        elif 'narration' in c_clean:
            narration_col = cols[c]

        elif 'credit' in c_clean:
            credit_col = cols[c]

    exclusion_keywords = [
        'lodging',
        'telephone',
        'postage',
        'stationery',
        'printing',
        'cartage',
        'freight',
        'transport',
        'motor car',
        'vehicle repair',
        'tyres',
        'tds',
        'cgst',
        'sgst',
        'igst',
        'gst',
        'round off',
        'cash discount',
        'advertising',
        'hotel',
        'repair charges'
    ]

    narration_exclusions = [
        'party merged',
        'balance transfer'
    ]

    for i in data_df.index:

        party = str(data_df.at[i, party_col]).lower() \
            if party_col else ''

        particulars = str(data_df.at[i, particulars_col]).lower() \
            if particulars_col else ''

        narration = str(data_df.at[i, narration_col]).lower() \
            if narration_col else ''

        credit_amt = 0

        if credit_col:

            try:
                credit_raw = str(
                    data_df.at[i, credit_col]
                ).replace(',', '')

                credit_amt = float(credit_raw)

            except:
                credit_amt = 0

        include_row = (
            ('jaishil' in party) and
            ('jaishil' not in particulars) and
            (credit_amt > 0)
        )

        exclude_row = False

        if any(k in particulars for k in exclusion_keywords):
            exclude_row = True

        if any(k in narration for k in narration_exclusions):
            exclude_row = True

        if not include_row or exclude_row:

            blank_row(cleaned_df, i)

            changes.append({
                'row': i + 1,
                'reason': 'JV row excluded'
            })

    return raw_df, cleaned_df, changes
# ============================================================
# HO EXP CLEANING
# ============================================================

def clean_ho_exp_file(file_path):

    raw_df = pd.read_excel(file_path, header=None)

    cleaned_df = raw_df.copy()

    changes = []

    total_row = None

    for i in range(len(raw_df)):

        row_text = " ".join(
            [
                str(x).lower()
                for x in raw_df.iloc[i].tolist()
                if pd.notna(x)
            ]
        )

        if "grand total" in row_text:

            total_row = i
            break

    if total_row is None:

        changes.append({
            "row": 0,
            "reason": "Grand Total row not found"
        })

        return raw_df, cleaned_df, changes

    numeric_columns = []

    for col in range(len(raw_df.columns)):

        count_numeric = 0

        for row in range(total_row):

            try:

                value = str(
                    raw_df.iat[row, col]
                ).replace(",", "")

                float(value)

                count_numeric += 1

            except:
                pass

        if count_numeric >= 3:

            numeric_columns.append(col)

    for col in numeric_columns:

        calculated_total = 0

        for row in range(total_row):

            try:

                value = str(
                    raw_df.iat[row, col]
                ).replace(",", "")

                calculated_total += float(value)

            except:
                pass

        try:

            grand_total_value = str(
                raw_df.iat[total_row, col]
            ).replace(",", "")

            grand_total_value = float(grand_total_value)

        except:

            continue

        if round(calculated_total, 2) != round(grand_total_value, 2):

            changes.append({
                "row": total_row + 1,
                "reason":
                f"Column {col + 1} total mismatch "
                f"(Calculated={calculated_total:.2f}, "
                f"Grand Total={grand_total_value:.2f})"
            })

    return raw_df, cleaned_df, changes
# ============================================================
# SUNDRY CREDITORS CLEANING
# ============================================================

def clean_sundry_creditors_file(file_path):

    raw_df = pd.read_excel(file_path, header=None)
    cleaned_df = raw_df.copy()

    changes = []

    company_name = "jaishil sulphur & chemical industries"

    for i in range(len(cleaned_df)):

        for j in range(len(cleaned_df.columns)):

            value = str(cleaned_df.iat[i, j]).strip()

            if value.lower() == company_name:

                cleaned_df.iat[i, j] = ""

                changes.append({
                    'row': i + 1,
                    'reason': 'Company name removed from creditors column'
                })

    return raw_df, cleaned_df, changes

# ============================================================
# SAVE WORKBOOK
# ============================================================

def find_header_row(df):
    header_keywords = {
        'si no.', 'si no', 'sl no.', 'sl no', 'sr no.', 'sr no', 'sr. no.', 'sr.no.', 
        's.no.', 's. no.', 'sno', 'sno.', 'particulars', 'particular', 'material name', 
        'customer name', 'partyname', 'party name', 'narration', 'inv no', 'inv no.', 
        'inv dt', 'date', 'debit', 'credit', 'amount', 'supplier', 'sundry creditors', 
        'group summary', 'opening balance', 'closing balance', 'grand total', 'total',
        'material'
    }
    for idx, row in df.iterrows():
        vals = [str(x).strip().lower() for x in row.tolist() if pd.notna(x)]
        for val in vals:
            if any(k in val for k in header_keywords):
                return idx
    return 0  # Fallback to index 0

def save_cleaned_workbook(
    raw_df,
    cleaned_df,
    changes,
    output_path,
    cleaned_sheet_name
):

    wb = Workbook()

    # RAW SHEET
    ws1 = wb.active
    ws1.title = 'Original_Raw'

    for r in raw_df.itertuples(index=False):
        # Convert nan values in raw sheet to None for a cleaner look, and format date-time
        raw_vals = [None if (pd.isna(x) or str(x).strip().lower() in ('nan', 'nat')) else format_date_val(x) for x in list(r)]
        ws1.append(raw_vals)

    # CLEANED SHEET
    ws2 = wb.create_sheet(cleaned_sheet_name)

    # Filter cleaned_df: drop metadata rows before the header row, and all entirely empty rows
    header_row_idx = find_header_row(cleaned_df)
    sub_df = cleaned_df.iloc[header_row_idx:].copy()

    def is_row_empty(row):
        return all(pd.isna(x) or str(x).strip() == '' or str(x).strip().lower() in ('nan', 'nat') for x in row)

    filtered_rows = []
    if len(sub_df) > 0:
        # Keep header row
        header_vals = [None if (pd.isna(x) or str(x).strip().lower() in ('nan', 'nat')) else format_date_val(x) for x in list(sub_df.iloc[0])]
        filtered_rows.append(header_vals)
        # Keep non-empty data rows and format date-time
        for i in range(1, len(sub_df)):
            r_vals = list(sub_df.iloc[i])
            if not is_row_empty(r_vals):
                cleaned_vals = [None if (pd.isna(x) or str(x).strip() == '' or str(x).strip().lower() in ('nan', 'nat')) else format_date_val(x) for x in r_vals]
                filtered_rows.append(cleaned_vals)

    for r in filtered_rows:
        ws2.append(r)

    # Style the header row (now Row 1)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", start_color="D9E1F2")
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border_side = Side(style='thin', color='A0A0A0')
    header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    max_cols = ws2.max_column
    for c in range(1, max_cols + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Adjust column widths dynamically based on headers
    for col_idx in range(1, max_cols + 1):
        col_letter = get_column_letter(col_idx)
        val = ws2.cell(row=1, column=col_idx).value
        val_str = str(val).lower() if val else ''
        if any(k in val_str for k in ('particular', 'material', 'customer', 'supplier', 'party', 'ledger', 'company', 'name')):
            ws2.column_dimensions[col_letter].width = 55
        else:
            ws2.column_dimensions[col_letter].width = 20

    # CHANGES LOG
    ws3 = wb.create_sheet('Changes_Log')

    headers = ['Row Number', 'Reason']
    log_header_font = Font(bold=True, size=11)
    log_header_fill = PatternFill("solid", start_color="FFE699")
    log_header_align = Alignment(horizontal='center')

    for col, h in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=col)
        cell.value = h
        cell.font = log_header_font
        cell.fill = log_header_fill
        cell.alignment = log_header_align
        cell.border = header_border

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 60

    for idx, change in enumerate(changes, start=2):
        ws3.cell(row=idx, column=1, value=change['row'])
        ws3.cell(row=idx, column=2, value=change['reason'])

    # Ensure grid lines are visible on all sheets
    for ws in (ws1, ws2, ws3):
        if ws.views.sheetView:
            ws.views.sheetView[0].showGridLines = True
        else:
            ws.sheet_view.showGridLines = True

    wb.save(output_path)

# ============================================================
# SUMMARY TXT
# ============================================================

def generate_summary_txt(
    filename,
    file_type,
    changes,
    output_path
):

    with open(output_path, "w", encoding="utf-8") as f:

        f.write("DATA CLEANING SUMMARY\n")
        f.write("=" * 50 + "\n\n")

        f.write(f"File Name : {filename}\n")
        f.write(f"Detected Type : {file_type}\n")
        f.write(f"Total Changes : {len(changes)}\n\n")

        f.write("CHANGES MADE:\n")
        f.write("-" * 50 + "\n")

        if len(changes) == 0:

            f.write("No changes were required.\n")

        else:

            for idx, change in enumerate(changes, start=1):

                f.write(
                    f"{idx}. Row {change['row']} → "
                    f"{change['reason']}\n"
                )


